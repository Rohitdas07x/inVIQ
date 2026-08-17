"""
Data Import Service — deterministic file parsing, confidence gating, batch processing, and quarantine management.

Guarantees:
  - NO AI INVOLVEMENT during row processing.
  - Stream-parses CSV and Excel files.
  - Applies confidence gating per field against configurable thresholds.
  - Commits in configurable batches (settings.IMPORT_BATCH_SIZE).
  - Quarantines failed or low-confidence rows to the DB table import_quarantine_rows.
  - Executes synchronously for files <= IMPORT_SYNC_ROW_LIMIT or asynchronously in background thread.
"""

import csv
import io
import logging
from datetime import date, datetime
from typing import Dict, Any, List, Tuple, Optional, Generator

from sqlalchemy.orm import Session

from app.core.config import settings
from app.infrastructure.database.models import DataImportJob, Location, Item, InventoryTransaction
from app.infrastructure.database.data_import_repo import DataImportRepository
from app.infrastructure.database.inventory_repo import InventoryRepository
from app.application.inventory_service import InventoryService
from app.application.data_import_mapper import DataImportMapper, HIGH_RISK_FIELDS
from app.application.cache_service import cache_invalidate_pattern

logger = logging.getLogger("smart_inventory.data_import_service")


class DataImportService:
    """Handles deterministic file parsing, mapping application, validation, and batch execution."""

    def __init__(self, db: Session):
        self.db = db
        self.import_repo = DataImportRepository(db)
        self.inv_repo = InventoryRepository(db)
        self.inv_service = InventoryService(self.inv_repo)
        self.mapper = DataImportMapper()

    # ── 1. File Inspection & Sample Extraction ────────────────────────────────

    @staticmethod
    def inspect_file(
        file_bytes: bytes,
        filename: str,
        sample_size: int = None,
    ) -> Tuple[List[str], List[Dict[str, Any]], int]:
        """
        Stream-reads file header, extracts a small sample, and counts data rows.
        Never loads entire dataset into memory.
        
        Returns: (headers, sample_rows, total_row_count)
        """
        if sample_size is None:
            sample_size = settings.IMPORT_SAMPLE_ROWS

        is_excel = filename.lower().endswith((".xlsx", ".xls"))

        if is_excel:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            if not ws:
                raise ValueError("Excel workbook has no active sheet")

            rows_iter = ws.iter_rows(values_only=True)
            try:
                first_row = next(rows_iter)
            except StopIteration:
                wb.close()
                return [], [], 0

            headers = [str(h).strip() for h in first_row if h is not None]
            sample_rows: List[Dict[str, Any]] = []
            total_count = 0

            for row in rows_iter:
                # Ignore empty trailing rows
                if not any(row):
                    continue
                total_count += 1
                if len(sample_rows) < sample_size:
                    row_dict = {}
                    for idx, h in enumerate(headers):
                        val = row[idx] if idx < len(row) else None
                        row_dict[h] = str(val) if val is not None else ""
                    sample_rows.append(row_dict)

            wb.close()
            return headers, sample_rows, total_count

        else:
            # CSV parser with encoding auto-detection fallback
            try:
                text = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text = file_bytes.decode("latin-1")

            stream = io.StringIO(text)
            # Sniff delimiter (comma, semicolon, tab)
            sample_chunk = stream.read(2048)
            stream.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample_chunk)
            except Exception:
                dialect = csv.excel

            reader = csv.reader(stream, dialect=dialect)
            try:
                first_row = next(reader)
            except StopIteration:
                return [], [], 0

            headers = [h.strip() for h in first_row if h.strip()]
            sample_rows = []
            total_count = 0

            for row in reader:
                if not any(row):
                    continue
                total_count += 1
                if len(sample_rows) < sample_size:
                    row_dict = {
                        headers[i]: row[i] if i < len(row) else ""
                        for i in range(len(headers))
                    }
                    sample_rows.append(row_dict)

            return headers, sample_rows, total_count

    # ── 2. Stream Row Generator ───────────────────────────────────────────────

    @staticmethod
    def _stream_rows(
        file_bytes: bytes,
        filename: str,
        headers: List[str],
    ) -> Generator[Tuple[int, Dict[str, Any]], None, None]:
        """Yield (row_number_1_indexed, raw_row_dict) one by one."""
        is_excel = filename.lower().endswith((".xlsx", ".xls"))

        if is_excel:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            next(rows_iter, None)  # Skip header

            for row_idx, row in enumerate(rows_iter, start=2):
                if not any(row):
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    val = row[idx] if idx < len(row) else None
                    row_dict[h] = val
                yield row_idx, row_dict

            wb.close()
        else:
            try:
                text = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text = file_bytes.decode("latin-1")

            stream = io.StringIO(text)
            try:
                dialect = csv.Sniffer().sniff(text[:2048])
            except Exception:
                dialect = csv.excel

            reader = csv.reader(stream, dialect=dialect)
            next(reader, None)  # Skip header

            for row_idx, row in enumerate(reader, start=2):
                if not any(row):
                    continue
                row_dict = {
                    headers[i]: row[i] if i < len(row) else None
                    for i in range(len(headers))
                }
                yield row_idx, row_dict

    # ── 3. Deterministic Validation & Confidence Gating ───────────────────────

    @staticmethod
    def _parse_date(val: Any) -> Optional[date]:
        """Convert various date representations to date object."""
        if not val:
            return None
        if isinstance(val, date):
            return val
        if isinstance(val, datetime):
            return val.date()
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    def transform_and_validate_row(
        self,
        raw_row: Dict[str, Any],
        mapping_result: Dict[str, Any],
        target_entity: str,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
        """
        Transforms raw row into target field dictionary based on confirmed mappings.
        Applies confidence gating and type coercion.

        Returns: (transformed_data, quarantine_info)
        If valid: (data_dict, None)
        If invalid: (None, {"reason": str, "field_name": str, "confidence_score": float})
        """
        mappings = mapping_result.get("mappings", {})
        high_risk_set = HIGH_RISK_FIELDS.get(target_entity, set())
        transformed = {}

        # 1. Map columns and check confidence gating
        for header, map_info in mappings.items():
            target_field = map_info.get("target_field")
            if not target_field:
                continue

            confidence = float(map_info.get("confidence", 1.0))
            required_threshold = (
                settings.IMPORT_HIGH_RISK_CONFIDENCE
                if target_field in high_risk_set
                else settings.IMPORT_DEFAULT_CONFIDENCE
            )

            if confidence < required_threshold:
                return None, {
                    "reason": "LOW_CONFIDENCE",
                    "field_name": target_field,
                    "confidence_score": confidence,
                }

            val = raw_row.get(header)
            if val is not None and str(val).strip() != "":
                transformed[target_field] = val

        # 2. Entity-specific field validation and type conversion
        if target_entity == "inventory_transaction":
            # Must have item_name or item_id
            if not transformed.get("item_name") and not transformed.get("item_id"):
                return None, {"reason": "MISSING_REQUIRED", "field_name": "item_name", "confidence_score": None}

            # Transaction date
            raw_dt = transformed.get("date")
            tx_date = self._parse_date(raw_dt) if raw_dt else date.today()
            if raw_dt and not tx_date:
                return None, {"reason": "VALIDATION_ERROR", "field_name": "date", "confidence_score": None}
            transformed["date"] = tx_date

            # Received / Issued quantities
            try:
                rec = int(transformed.get("received", 0) or 0)
                iss = int(transformed.get("issued", 0) or 0)
                if rec < 0 or iss < 0:
                    return None, {"reason": "VALIDATION_ERROR", "field_name": "received/issued", "confidence_score": None}
                if rec == 0 and iss == 0:
                    # If neither received nor issued explicitly given, default to received=1
                    rec = 1
                transformed["received"] = rec
                transformed["issued"] = iss
            except (ValueError, TypeError):
                return None, {"reason": "VALIDATION_ERROR", "field_name": "received/issued", "confidence_score": None}

            # Optional batch & expiry
            if "expiry_date" in transformed and transformed["expiry_date"]:
                exp_dt = self._parse_date(transformed["expiry_date"])
                if not exp_dt:
                    return None, {"reason": "VALIDATION_ERROR", "field_name": "expiry_date", "confidence_score": None}
                transformed["expiry_date"] = exp_dt

            if "batch_number" in transformed:
                transformed["batch_number"] = str(transformed["batch_number"]).strip()

        elif target_entity == "item":
            if not transformed.get("name") or not str(transformed["name"]).strip():
                return None, {"reason": "MISSING_REQUIRED", "field_name": "name", "confidence_score": None}

            transformed["name"] = str(transformed["name"]).strip()
            transformed["category"] = str(transformed.get("category", "general")).strip().lower()
            transformed["unit"] = str(transformed.get("unit", "units")).strip().lower()

            try:
                transformed["lead_time_days"] = int(transformed.get("lead_time_days", 7) or 7)
                transformed["min_stock"] = int(transformed.get("min_stock", 10) or 10)
            except (ValueError, TypeError):
                return None, {"reason": "VALIDATION_ERROR", "field_name": "lead_time_days/min_stock", "confidence_score": None}

            storage = str(transformed.get("storage_temp", "ambient")).strip().lower()
            transformed["storage_temp"] = "cold_chain" if "cold" in storage else "ambient"

        elif target_entity == "location":
            if not transformed.get("name") or not str(transformed["name"]).strip():
                return None, {"reason": "MISSING_REQUIRED", "field_name": "name", "confidence_score": None}

            transformed["name"] = str(transformed["name"]).strip()
            transformed["type"] = str(transformed.get("type", "warehouse")).strip().lower()
            transformed["region"] = str(transformed.get("region", "Default")).strip()
            transformed["address"] = str(transformed.get("address", "")).strip() or None

        return transformed, None

    # ── 4. Batch Execution Engine ─────────────────────────────────────────────

    def execute_import(
        self,
        job_id: int,
        confirmed_mapping: Dict[str, Any],
        default_location_id: Optional[int] = None,
        entered_by: str = "import_system",
    ) -> DataImportJob:
        """
        Executes the import job in stream-processing batches.
        Updates job state, writes valid records to target tables,
        and logs rejected rows to quarantine.
        """
        job = self.import_repo.get_job(job_id)
        if not job:
            raise ValueError(f"Job #{job_id} not found")

        if not job.file_content:
            job.status = "FAILED"
            job.error_message = "File content is missing from job record"
            return self.import_repo.update_job(job)

        job.status = "PROCESSING"
        self.import_repo.update_job(job)

        # Pre-cache existing items/locations for quick matching during transaction imports
        item_cache: Dict[str, Item] = {}
        location_cache: Dict[str, Location] = {}
        if job.target_entity == "inventory_transaction":
            for it in self.inv_repo.get_all_items(org_id=job.org_id):
                item_cache[it.name.lower()] = it
            for loc in self.inv_repo.get_all_locations(org_id=job.org_id):
                location_cache[loc.name.lower()] = loc

        success_count = 0
        quarantine_count = 0
        batch_size = settings.IMPORT_BATCH_SIZE

        quarantine_buffer: List[Dict[str, Any]] = []
        row_generator = self._stream_rows(
            file_bytes=job.file_content,
            filename=job.filename,
            headers=list(confirmed_mapping.get("mappings", {}).keys()),
        )

        try:
            for row_number, raw_row in row_generator:
                transformed, quarantine_info = self.transform_and_validate_row(
                    raw_row=raw_row,
                    mapping_result=confirmed_mapping,
                    target_entity=job.target_entity,
                )

                if quarantine_info:
                    quarantine_buffer.append({
                        "job_id": job.id,
                        "row_number": row_number,
                        "raw_data": raw_row,
                        "reason": quarantine_info["reason"],
                        "field_name": quarantine_info.get("field_name"),
                        "confidence_score": quarantine_info.get("confidence_score"),
                    })
                    quarantine_count += 1
                else:
                    # Write to domain/repository
                    try:
                        self._write_single_entity(
                            job=job,
                            data=transformed,
                            default_location_id=default_location_id,
                            entered_by=entered_by,
                            item_cache=item_cache,
                            location_cache=location_cache,
                        )
                        success_count += 1
                    except Exception as err:
                        quarantine_buffer.append({
                            "job_id": job.id,
                            "row_number": row_number,
                            "raw_data": raw_row,
                            "reason": "DB_ERROR",
                            "field_name": str(err)[:100],
                            "confidence_score": None,
                        })
                        quarantine_count += 1

                # Commit batch
                if (success_count + quarantine_count) % batch_size == 0:
                    if quarantine_buffer:
                        self.import_repo.add_quarantine_rows_bulk(quarantine_buffer)
                        quarantine_buffer.clear()
                    self.db.commit()

            # Flush final batch
            if quarantine_buffer:
                self.import_repo.add_quarantine_rows_bulk(quarantine_buffer)
                quarantine_buffer.clear()
            self.db.commit()

            # Invalidate cache
            cache_invalidate_pattern("analytics:*")

            # Final status
            job.success_rows = success_count
            job.quarantined_rows = quarantine_count
            job.total_rows = success_count + quarantine_count
            job.status = (
                "COMPLETED" if quarantine_count == 0
                else "PARTIAL" if success_count > 0
                else "FAILED"
            )

        except Exception as e:
            self.db.rollback()
            logger.error("Data import job #%d encountered fatal error: %s", job.id, str(e))
            job.status = "FAILED"
            job.error_message = str(e)
            job.success_rows = success_count
            job.quarantined_rows = quarantine_count

        return self.import_repo.update_job(job)

    def _write_single_entity(
        self,
        job: DataImportJob,
        data: Dict[str, Any],
        default_location_id: Optional[int],
        entered_by: str,
        item_cache: Dict[str, Item],
        location_cache: Dict[str, Location],
    ) -> None:
        """Write single transformed row to repository based on target_entity."""
        target = job.target_entity

        if target == "inventory_transaction":
            # Match Item
            item_id = data.get("item_id")
            if not item_id:
                item_name = str(data.get("item_name", "")).strip()
                matched_item = item_cache.get(item_name.lower())
                if not matched_item:
                    # Auto-create item within caller organization if it doesn't exist
                    matched_item = self.inv_repo.create_item(
                        name=item_name,
                        category="general",
                        unit="units",
                        lead_time_days=7,
                        min_stock=10,
                        storage_temp="ambient",
                        org_id=job.org_id,
                    )
                    item_cache[item_name.lower()] = matched_item
                item_id = matched_item.id

            # Match Location
            location_id = data.get("location_id") or default_location_id
            if not location_id and data.get("location_name"):
                loc_name = str(data["location_name"]).strip()
                matched_loc = location_cache.get(loc_name.lower())
                if not matched_loc:
                    matched_loc = self.inv_repo.create_location(
                        name=loc_name,
                        type="warehouse",
                        region="Default",
                        org_id=job.org_id,
                    )
                    location_cache[loc_name.lower()] = matched_loc
                location_id = matched_loc.id

            if not location_id:
                # Fallback to first available location in caller organization
                all_locs = self.inv_repo.get_all_locations(org_id=job.org_id)
                if all_locs:
                    location_id = all_locs[0].id
                else:
                    new_loc = self.inv_repo.create_location(
                        name="Main Facility",
                        type="warehouse",
                        region="Default",
                        org_id=job.org_id,
                    )
                    location_id = new_loc.id

            self.inv_service.add_transaction(
                location_id=location_id,
                item_id=item_id,
                transaction_date=data["date"],
                received=data.get("received", 0),
                issued=data.get("issued", 0),
                notes=data.get("notes") or f"Imported via {job.filename}",
                entered_by=entered_by,
                flush_only=True,
                batch_number=data.get("batch_number"),
                expiry_date=data.get("expiry_date"),
            )

        elif target == "item":
            existing = self.inv_repo.get_item_by_name(data["name"], org_id=job.org_id)
            if existing:
                existing.category = data.get("category", existing.category)
                existing.unit = data.get("unit", existing.unit)
                existing.lead_time_days = data.get("lead_time_days", existing.lead_time_days)
                existing.min_stock = data.get("min_stock", existing.min_stock)
                existing.storage_temp = data.get("storage_temp", existing.storage_temp)
                self.db.add(existing)
            else:
                new_item = Item(
                    org_id=job.org_id,
                    name=data["name"],
                    category=data.get("category", "general"),
                    unit=data.get("unit", "units"),
                    lead_time_days=data.get("lead_time_days", 7),
                    min_stock=data.get("min_stock", 10),
                    storage_temp=data.get("storage_temp", "ambient"),
                )
                self.db.add(new_item)

        elif target == "location":
            existing = self.inv_repo.get_location_by_name(data["name"], org_id=job.org_id)
            if existing:
                existing.type = data.get("type", existing.type)
                existing.region = data.get("region", existing.region)
                existing.address = data.get("address", existing.address)
                self.db.add(existing)
            else:
                new_loc = Location(
                    org_id=job.org_id,
                    name=data["name"],
                    type=data.get("type", "warehouse"),
                    region=data.get("region", "Default"),
                    address=data.get("address"),
                )
                self.db.add(new_loc)

